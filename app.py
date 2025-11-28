import streamlit as st
import pandas as pd
import numpy as np
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation


st.set_page_config(page_title="Compensaciones por Tiempo en Losa", layout="wide")
st.title("üì¶ Compensaciones por Tiempo en Losa ‚Äî Cabify Style")


# ------------------------------
# Funci√≥n para calcular compensaci√≥n
# ------------------------------
def calcular_compensacion(minutos):
    if pd.isna(minutos):
        return 9000
    try:
        minutos = float(minutos)
    except:
        return 9000

    if minutos >= 50:
        return 9000
    elif minutos >= 40:
        return 6000
    elif minutos >= 35:
        return 3000
    else:
        return 0


# ------------------------------
# Subir archivo CSV
# ------------------------------
uploaded_file = st.file_uploader("üì§ Sube tu archivo CSV", type=["csv"])

if uploaded_file is not None:

    try:
        df = pd.read_csv(uploaded_file, dtype=str)
        st.success("Archivo cargado correctamente üéâ")
    except Exception as e:
        st.error(f"‚ùå Error al leer el CSV: {e}")
        st.stop()

    # Columnas necesarias (ahora incluye User Email)
    columnas = [
        "Day of tm_start_local_at",
        "Segmento Tiempo en Losa",
        "End State",
        "id_reservation_id",
        "Service Channel",
        "Minutes Creation - Pickup",
        "User Fullname",
        "User Email",
        "User Phone Number"
    ]

    faltantes = [c for c in columnas if c not in df.columns]
    if faltantes:
        st.error(f"‚ùå Faltan columnas requeridas: {faltantes}")
        st.stop()

    df = df[columnas].copy()

    # Convertir fecha
    df["Day of tm_start_local_at"] = pd.to_datetime(
        df["Day of tm_start_local_at"], errors="coerce"
    ).dt.date

    if df["Day of tm_start_local_at"].isna().all():
        st.error("‚ùå No hay fechas v√°lidas.")
        st.stop()

    # ------------------------------
    # Filtro de fechas
    # ------------------------------
    fecha_min = df["Day of tm_start_local_at"].min()
    fecha_max = df["Day of tm_start_local_at"].max()

    fecha_desde, fecha_hasta = st.date_input(
        "üìÖ Selecciona rango de fechas:",
        value=(fecha_min, fecha_max)
    )

    df = df[
        (df["Day of tm_start_local_at"] >= fecha_desde) &
        (df["Day of tm_start_local_at"] <= fecha_hasta)
    ]

    if df.empty:
        st.warning("‚ö†Ô∏è No hay registros en ese rango.")
        st.stop()

    # Estado Pago por defecto = No Pagado
    df["Estado Pago"] = "No Pagado"

    # ------------------------------
    # C√°lculo de compensaci√≥n
    # ------------------------------
    df["Minutes Creation - Pickup"] = pd.to_numeric(
        df["Minutes Creation - Pickup"], errors="coerce"
    )

    df["Monto a Reembolsar"] = df["Minutes Creation - Pickup"].apply(calcular_compensacion)

    df = df[df["Monto a Reembolsar"] > 0]

    if df.empty:
        st.warning("‚ö†Ô∏è No hay compensaciones > 0.")
        st.stop()

    # ------------------------------
    # Mostrar datos filtrados
    # ------------------------------
    st.subheader("üìä Registros procesados")
    st.dataframe(df, use_container_width=True)

    # ------------------------------
    # RESUMEN POR TRAMO DE COMPENSACI√ìN
    # ------------------------------
    st.subheader("üìà Resumen de compensaciones")

    resumen = df["Monto a Reembolsar"].value_counts().sort_index()
    total_casos = len(df)
    total_dinero = df["Monto a Reembolsar"].sum()

    st.write("### üßÆ Cantidad de casos por monto:")
    for monto, cantidad in resumen.items():
        st.write(f"- **${monto:,}** ‚Üí {cantidad} casos")

    st.write("### üí∞ Total compensaciones:")
    st.write(f"- **Total de casos:** {total_casos}")
    st.write(f"- **Total en dinero:** ${total_dinero:,}")

    # ------------------------------
    # CREAR EXCEL ESTILIZADO CABIFY
    # ------------------------------
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "Compensaciones"

    # Estilos Cabify
    header_fill = PatternFill("solid", fgColor="5B34AC")  # Morado Cabify
    header_font = Font(color="FFFFFF", bold=True)
    border_side = Side(style="thin", color="362065")
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)

    color_no_pagado = PatternFill("solid", fgColor="EA8C2E")
    color_pagado = PatternFill("solid", fgColor="0C936B")

    fill_3000 = PatternFill("solid", fgColor="EFBD03")
    fill_6000 = PatternFill("solid", fgColor="EA8C2E")
    fill_9000 = PatternFill("solid", fgColor="E83C96")

    alt_fill = PatternFill("solid", fgColor="FAF8FE")

    # Escribir headers + datos
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)

    # Encabezado
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Activar filtros
    ws.auto_filter.ref = f"A1:{chr(64 + len(df.columns))}1"

    # Estilos por fila
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        idx = row[0].row

        # Filas alternadas
        if idx % 2 == 0:
            for cell in row:
                cell.fill = alt_fill

        # Color Estado Pago
        estado = row[df.columns.get_loc("Estado Pago")].value
        if estado == "No Pagado":
            row[df.columns.get_loc("Estado Pago")].fill = color_no_pagado
        else:
            row[df.columns.get_loc("Estado Pago")].fill = color_pagado

        # Color compensaci√≥n
        monto = float(row[df.columns.get_loc("Monto a Reembolsar")].value)
        if monto == 3000:
            row[df.columns.get_loc("Monto a Reembolsar")].fill = fill_3000
        elif monto == 6000:
            row[df.columns.get_loc("Monto a Reembolsar")].fill = fill_6000
        elif monto == 9000:
            row[df.columns.get_loc("Monto a Reembolsar")].fill = fill_9000

        # Bordes
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(vertical="center")

    # ------------------------------
    # COMBOBOX Estado Pago
    # ------------------------------
    dv = DataValidation(type="list", formula1='"Pagado,No Pagado"', allow_blank=False)
    col_estado_pago = df.columns.get_loc("Estado Pago") + 1
    col_letter = chr(64 + col_estado_pago)
    dv.add(f"{col_letter}2:{col_letter}1048576")
    ws.add_data_validation(dv)

    # Guardar Excel
    wb.save(output)
    output.seek(0)

    # Descargar
    st.download_button(
        "‚¨áÔ∏è Descargar Excel con Estilo Cabify",
        output,
        file_name="compensaciones_losa_cabify.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

else:
    st.info("Sube un archivo CSV para comenzar.")


